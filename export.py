from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
from decimal import Decimal
from typing import Optional
import os

from classes import (
    Estimate, Section, WorkItem, Resource, AbstractResource,
    SectionSummary, RowType, ResourceType
)


class EstimateExcelExporter:
    """Экспорт сметы в XLSX точно по шаблону ФГИС ЦС"""

    # Ширина колонок
    COLUMN_WIDTHS = {
        1: 5,    # № п/п
        2: 30,   # Обоснование
        3: 50,   # Наименование
        4: 12,   # Ед. изм.
        5: 12,   # Количество на ед.
        6: 12,   # Коэффициенты
        7: 15,   # Всего с коэфф.
        8: 15,   # Цена базисная
        9: 10,   # Индекс
        10: 15,  # Цена текущая
        11: 12,  # Коэффициент доп.
        12: 18,  # Всего стоимость
    }

    # Стили
    STYLE_HEADER_MAIN = {
        'font': Font(bold=True, size=11, name='Calibri'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        ),
        'fill': PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid'),
    }

    STYLE_HEADER_SUB = {
        'font': Font(bold=True, size=10, name='Calibri'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        ),
    }

    STYLE_SECTION = {
        'font': Font(bold=True, size=11, name='Calibri'),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'fill': PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid'),
    }

    STYLE_WORK_MAIN = {
        'font': Font(bold=True, size=10, name='Calibri'),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'border': Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        ),
    }

    STYLE_COEFFICIENT = {
        'font': Font(italic=True, size=9, name='Calibri'),
        'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'border': Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        ),
    }

    STYLE_RESOURCE = {
        'font': Font(size=9, name='Calibri'),
        'alignment': Alignment(horizontal='right', vertical='center'),
        'border': Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        ),
    }

    STYLE_REGULAR = {
        'font': Font(size=9, name='Calibri'),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'border': Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        ),
    }

    STYLE_SUBTOTAL = {
        'font': Font(bold=True, size=10, name='Calibri'),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'border': Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        ),
    }

    STYLE_TOTAL = {
        'font': Font(bold=True, size=11, name='Calibri'),
        'alignment': Alignment(horizontal='right', vertical='center'),
        'border': Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='medium'), bottom=Side(style='medium')
        ),
        'fill': PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid'),
    }

    def __init__(self, estimate: Estimate):
        self.estimate = estimate
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Лист1"
        self.current_row = 1
        self.current_item_num = 1

    def _apply_style(self, cell, style: dict) -> None:
        for key, value in style.items():
            setattr(cell, key, value)

    def _merge_cells(self, min_col: int, max_col: int, min_row: int, max_row: int) -> None:
        self.ws.merge_cells(
            start_row=min_row, end_row=max_row,
            start_column=min_col, end_column=max_col
        )

    def _format_number(self, value: Decimal) -> Optional[float]:
        """Форматирует число, НЕ округляет. Если число равно 1, возвращает None."""
        if value is None:
            return None
        if abs(float(value) - 1.0) < 0.000001:
            return None
        return float(value)

    def _write_metadata_header(self) -> None:
        """Шапка с метаданными (строки 1-25)"""
        meta = self.estimate.metadata

        self._merge_cells(1, 12, self.current_row, self.current_row)
        cell = self.ws.cell(row=self.current_row, column=1, value=meta.software_name)
        self._apply_style(cell, {
            'font': Font(bold=True, size=12, name='Calibri'),
            'alignment': Alignment(horizontal='center', vertical='center'),
        })
        self.current_row += 1

        self._merge_cells(1, 5, self.current_row, self.current_row + 3)
        cell = self.ws.cell(row=self.current_row, column=1,
                           value="Наименование редакции сметных нормативов")
        self._apply_style(cell, {
            'font': Font(bold=True, size=10, name='Calibri'),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        })

        editions_text = ";\n".join(meta.normative_editions)
        self._merge_cells(6, 12, self.current_row, self.current_row + 3)
        cell = self.ws.cell(row=self.current_row, column=6, value=editions_text)
        self._apply_style(cell, {
            'font': Font(size=9, name='Calibri'),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
            'border': Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            ),
        })
        self.current_row += 4

        self._merge_cells(1, 5, self.current_row, self.current_row + 3)
        cell = self.ws.cell(row=self.current_row, column=1,
                           value="Реквизиты приказа Минстроя России об утверждении дополнений и изменений к сметным нормативам")
        self._apply_style(cell, {
            'font': Font(bold=True, size=10, name='Calibri'),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        })

        amendments_text = ";\n".join(meta.amendments)
        self._merge_cells(6, 12, self.current_row, self.current_row + 3)
        cell = self.ws.cell(row=self.current_row, column=6, value=amendments_text)
        self._apply_style(cell, {
            'font': Font(size=9, name='Calibri'),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
            'border': Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            ),
        })
        self.current_row += 4

        self._merge_cells(1, 5, self.current_row, self.current_row + 1)
        cell = self.ws.cell(row=self.current_row, column=1,
                           value="Реквизиты письма Минстроя России об индексах изменения сметной стоимости строительства")
        self._apply_style(cell, {
            'font': Font(bold=True, size=10, name='Calibri'),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        })

        self._merge_cells(6, 12, self.current_row, self.current_row + 1)
        cell = self.ws.cell(row=self.current_row, column=6, value=meta.indices_letter)
        self._apply_style(cell, {
            'font': Font(size=9, name='Calibri'),
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            ),
        })
        self.current_row += 2

        self._merge_cells(1, 5, self.current_row, self.current_row + 1)
        cell = self.ws.cell(row=self.current_row, column=1,
                           value="Реквизиты нормативного правового акта об утверждении оплаты труда")
        self._apply_style(cell, {
            'font': Font(bold=True, size=10, name='Calibri'),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        })

        self._merge_cells(6, 12, self.current_row, self.current_row + 1)
        cell = self.ws.cell(row=self.current_row, column=6, value=meta.labor_act)
        self._apply_style(cell, {
            'font': Font(size=9, name='Calibri'),
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            ),
        })
        self.current_row += 2

        self._merge_cells(1, 5, self.current_row, self.current_row)
        cell = self.ws.cell(row=self.current_row, column=1,
                           value="Наименование субъекта Российской Федерации")
        self._apply_style(cell, {
            'font': Font(bold=True, size=10, name='Calibri'),
            'alignment': Alignment(horizontal='left', vertical='center'),
        })

        self._merge_cells(6, 12, self.current_row, self.current_row)
        cell = self.ws.cell(row=self.current_row, column=6, value=meta.region)
        self._apply_style(cell, {
            'font': Font(size=9, name='Calibri'),
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            ),
        })
        self.current_row += 1

        self._merge_cells(1, 5, self.current_row, self.current_row)
        cell = self.ws.cell(row=self.current_row, column=1,
                           value="Наименование зоны субъекта Российской Федерации")
        self._apply_style(cell, {
            'font': Font(bold=True, size=10, name='Calibri'),
            'alignment': Alignment(horizontal='left', vertical='center'),
        })

        self._merge_cells(6, 12, self.current_row, self.current_row)
        cell = self.ws.cell(row=self.current_row, column=6, value=meta.zone)
        self._apply_style(cell, {
            'font': Font(size=9, name='Calibri'),
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            ),
        })
        self.current_row += 1
        self.current_row += 1

        self._merge_cells(1, 12, self.current_row, self.current_row)
        cell = self.ws.cell(row=self.current_row, column=1, value=meta.project_address)
        self._apply_style(cell, {
            'font': Font(bold=True, size=11, name='Calibri'),
            'alignment': Alignment(horizontal='left', vertical='center'),
        })
        self.current_row += 1

        self._merge_cells(2, 11, self.current_row, self.current_row)
        cell = self.ws.cell(row=self.current_row, column=2, value=meta.project_name)
        self._apply_style(cell, {
            'font': Font(italic=True, size=10, name='Calibri'),
            'alignment': Alignment(horizontal='center', vertical='center'),
        })
        self.current_row += 1
        self.current_row += 1

        self._merge_cells(1, 12, self.current_row, self.current_row)
        cell = self.ws.cell(row=self.current_row, column=1, value=meta.object_name)
        self._apply_style(cell, {
            'font': Font(bold=True, size=11, name='Calibri'),
            'alignment': Alignment(horizontal='left', vertical='center'),
        })
        self.current_row += 1

        self._merge_cells(2, 11, self.current_row, self.current_row)
        cell = self.ws.cell(row=self.current_row, column=2, value=meta.object_detail)
        self._apply_style(cell, {
            'font': Font(italic=True, size=10, name='Calibri'),
            'alignment': Alignment(horizontal='center', vertical='center'),
        })
        self.current_row += 2

        self._merge_cells(2, 11, self.current_row, self.current_row)
        cell = self.ws.cell(row=self.current_row, column=2,
                           value=f"ЛОКАЛЬНЫЙ СМЕТНЫЙ РАСЧЕТ (СМЕТА) № {meta.estimate_number}")
        self._apply_style(cell, {
            'font': Font(bold=True, size=12, name='Calibri'),
            'alignment': Alignment(horizontal='center', vertical='center'),
        })
        self.current_row += 1
        self.current_row += 1

        self._merge_cells(2, 12, self.current_row, self.current_row)
        cell = self.ws.cell(row=self.current_row, column=2, value=meta.estimate_title)
        self._apply_style(cell, {
            'font': Font(bold=True, size=11, name='Calibri'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        })
        self.current_row += 1

        self._merge_cells(2, 11, self.current_row, self.current_row)
        cell = self.ws.cell(row=self.current_row, column=2, value=meta.work_description)
        self._apply_style(cell, {
            'font': Font(italic=True, size=10, name='Calibri'),
            'alignment': Alignment(horizontal='center', vertical='center'),
        })
        self.current_row += 2

        self.ws.cell(row=self.current_row, column=1, value="Составлен")
        self.ws.cell(row=self.current_row, column=3, value=meta.method)
        self.ws.cell(row=self.current_row, column=4, value="методом")
        self.current_row += 2

        self.ws.cell(row=self.current_row, column=1, value="Основание")
        self._merge_cells(3, 12, self.current_row, self.current_row)
        cell = self.ws.cell(row=self.current_row, column=3, value=meta.basis)
        self._apply_style(cell, {
            'font': Font(size=9, name='Calibri'),
            'alignment': Alignment(horizontal='left', vertical='center'),
        })
        self.current_row += 1

        self._merge_cells(3, 12, self.current_row, self.current_row)
        cell = self.ws.cell(row=self.current_row, column=3, value=meta.basis_description)
        self._apply_style(cell, {
            'font': Font(italic=True, size=9, name='Calibri'),
            'alignment': Alignment(horizontal='left', vertical='center'),
        })
        self.current_row += 2

        self.ws.cell(row=self.current_row, column=1, value="Составлен(а) в текущем (базисном) уровне цен")
        self._merge_cells(4, 12, self.current_row, self.current_row)
        cell = self.ws.cell(row=self.current_row, column=4,
                           value=f"{meta.price_level_current}     ({meta.price_level_base})")
        self._apply_style(cell, {
            'font': Font(bold=True, size=10, name='Calibri'),
            'alignment': Alignment(horizontal='left', vertical='center'),
        })
        self.current_row += 2

    def _write_summary_block(self) -> None:
        """Блок сводных показателей"""
        summary = self.estimate.summary

        self.ws.cell(row=self.current_row, column=1, value="Сметная стоимость")
        self._merge_cells(4, 5, self.current_row, self.current_row)
        cell = self.ws.cell(row=self.current_row, column=4, value="#REF!")
        self._apply_style(cell, {'font': Font(bold=True, size=10, name='Calibri')})
        self.ws.cell(row=self.current_row, column=6, value="тыс. руб.")

        self.ws.cell(row=self.current_row, column=7, value="Средства на оплату труда рабочих")
        self._merge_cells(10, 11, self.current_row, self.current_row)
        cell = self.ws.cell(row=self.current_row, column=10, value="#REF!")
        self._apply_style(cell, {'font': Font(bold=True, size=10, name='Calibri')})
        self.ws.cell(row=self.current_row, column=12, value="тыс.руб.")
        self.current_row += 1

        self.ws.cell(row=self.current_row, column=2, value="в том числе:")
        self.ws.cell(row=self.current_row, column=7, value="Нормативные затраты труда рабочих")
        self._merge_cells(10, 11, self.current_row, self.current_row)
        cell = self.ws.cell(row=self.current_row, column=10, value="#REF!")
        self._apply_style(cell, {'font': Font(bold=True, size=10, name='Calibri')})
        self.ws.cell(row=self.current_row, column=12, value="чел.-ч")
        self.current_row += 1

        self.ws.cell(row=self.current_row, column=2, value="строительных работ")
        self._merge_cells(4, 5, self.current_row, self.current_row)
        cell = self.ws.cell(row=self.current_row, column=4, value="#REF!")
        self._apply_style(cell, {'font': Font(size=9, name='Calibri')})
        self.ws.cell(row=self.current_row, column=6, value="тыс.руб.")

        self.ws.cell(row=self.current_row, column=7, value="Нормативные затраты труда машинистов")
        self._merge_cells(10, 11, self.current_row, self.current_row)
        cell = self.ws.cell(row=self.current_row, column=10, value="#REF!")
        self._apply_style(cell, {'font': Font(bold=True, size=10, name='Calibri')})
        self.ws.cell(row=self.current_row, column=12, value="чел.-ч")
        self.current_row += 1

        self.ws.cell(row=self.current_row, column=2, value="монтажных работ")
        self._merge_cells(4, 5, self.current_row, self.current_row)
        cell = self.ws.cell(row=self.current_row, column=4, value="#REF!")
        self._apply_style(cell, {'font': Font(size=9, name='Calibri')})
        self.ws.cell(row=self.current_row, column=6, value="тыс.руб.")
        self.current_row += 1

        self.ws.cell(row=self.current_row, column=2, value="оборудования")
        self._merge_cells(4, 5, self.current_row, self.current_row)
        cell = self.ws.cell(row=self.current_row, column=4, value="#REF!")
        self._apply_style(cell, {'font': Font(size=9, name='Calibri')})
        self.ws.cell(row=self.current_row, column=6, value="тыс.руб.")
        self.current_row += 1

        self.ws.cell(row=self.current_row, column=2, value="прочих затрат")
        self._merge_cells(4, 5, self.current_row, self.current_row)
        cell = self.ws.cell(row=self.current_row, column=4, value="#REF!")
        self._apply_style(cell, {'font': Font(size=9, name='Calibri')})
        self.ws.cell(row=self.current_row, column=6, value="тыс.руб.")
        self.current_row += 2

    def _write_table_header(self) -> None:
        """Заголовки таблицы (3 строки)"""
        headers_main = [
            ("№ п/п", 1, 1),
            ("Обоснование", 1, 1),
            ("Наименование работ и затрат", 1, 1),
            ("Единица измерения", 1, 1),
            ("Количество", 1, 3),
            ("Сметная стоимость, руб", 1, 5),
        ]

        col = 1
        for name, row_span, col_span in headers_main:
            if col_span > 1:
                self._merge_cells(col, col + col_span - 1, self.current_row, self.current_row)
            cell = self.ws.cell(row=self.current_row, column=col, value=name)
            self._apply_style(cell, self.STYLE_HEADER_MAIN)
            col += col_span

        self.current_row += 1

        headers_sub = [
            "", "", "", "",
            "на единицу измерения",
            "коэффициенты",
            "всего с учетом коэффициентов",
            "на единицу измерения\nв базисном уровне цен",
            "индекс",
            "на единицу измерения\nв текущем уровне цен",
            "коэффициент",
            "всего в текущем\nуровне цен",
        ]

        for col, name in enumerate(headers_sub, 1):
            cell = self.ws.cell(row=self.current_row, column=col, value=name)
            self._apply_style(cell, self.STYLE_HEADER_SUB)

        self.current_row += 1

        for col in range(1, 13):
            cell = self.ws.cell(row=self.current_row, column=col, value=str(col))
            self._apply_style(cell, {
                'font': Font(bold=True, size=9, name='Calibri'),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                ),
            })

        self.current_row += 1

    def _write_section(self, section: Section) -> None:
        """Запись раздела сметы"""
        # Заголовок раздела
        self._merge_cells(1, 12, self.current_row, self.current_row)
        cell = self.ws.cell(row=self.current_row, column=1, value=section.full_name)
        self._apply_style(cell, self.STYLE_SECTION)
        self.current_row += 1

        for item in section.items:
            item.row_num = self.current_item_num
            self.current_item_num += 1

            self._write_work_item(item, item.conversion_string)
            self._write_work_item_totals(item)

            for abs_res in item.abstract_resources:
                abs_res.row_num = self.current_item_num
                self.current_item_num += 1
                self._write_abstract_resource_item(abs_res)

    def _write_work_item(self, item: WorkItem, conversion_string: str = None) -> None:
            """Запись позиции ГЭСН"""

            # Строка с ГЭСН (основная)
            self.ws.cell(row=self.current_row, column=1, value=item.row_num)
            self.ws.cell(row=self.current_row, column=2, value=item.code)
            self.ws.cell(row=self.current_row, column=3, value=item.name)
            self.ws.cell(row=self.current_row, column=4, value=item.unit)
            self.ws.cell(row=self.current_row, column=5, value=float(item.quantity))
            self.ws.cell(row=self.current_row, column=6, value="1")
            self.ws.cell(row=self.current_row, column=7, value=float(item.quantity))

            for col in range(1, 8):
                self._apply_style(self.ws.cell(row=self.current_row, column=col), self.STYLE_WORK_MAIN)
            self.current_row += 1

            # ✅ Строка с объёмом работ (если есть)
            if item.conversion_string:
                cell = self.ws.cell(row=self.current_row, column=3, value=f"Объем={item.conversion_string}")
                self._apply_style(cell, self.STYLE_REGULAR)
                self.current_row += 1

            # ✅ ВСЕ коэффициенты (должны быть до разбора ресурсов)
            for coef in item.coefficients:
                cell = self.ws.cell(row=self.current_row, column=2, value=coef.code)
                self._apply_style(cell, self.STYLE_RESOURCE)
                self._merge_cells(3, 12, self.current_row, self.current_row)
                cell = self.ws.cell(row=self.current_row, column=3, value=coef.description)
                self._apply_style(cell, self.STYLE_REGULAR)
                self.current_row += 1

            # ✅ ПОСЛЕ коэффициентов начинаем разбор ресурсов
            # Группа ОТ (1. ОТ (ЗТ))
            labor_resources = [r for r in item.resources if r.resource_type == ResourceType.LABOR]

            self.ws.cell(row=self.current_row, column=2, value="1")
            self.ws.cell(row=self.current_row, column=3, value="ОТ (ЗТ)")
            self.ws.cell(row=self.current_row, column=4, value="чел.-ч")
            group_total_ot = sum(r.total_cost for r in labor_resources)
            self.ws.cell(row=self.current_row, column=12, value=float(group_total_ot))
            self._apply_style(self.ws.cell(row=self.current_row, column=2), self.STYLE_RESOURCE)
            self._apply_style(self.ws.cell(row=self.current_row, column=12), self.STYLE_RESOURCE)
            self.current_row += 1

            for res in labor_resources:
                self._write_resource_detail(res)

            # Группа ЭМ и ОТм (2. ЭМ и ОТм(ЗТм))
            equipment_resources = [r for r in item.resources if r.resource_type == ResourceType.EQUIPMENT]
            operator_resources = [r for r in item.resources if r.resource_type == ResourceType.LABOR_OPERATOR]

            self.ws.cell(row=self.current_row, column=2, value="2")
            self.ws.cell(row=self.current_row, column=3, value="ЭМ")
            self.ws.cell(row=self.current_row, column=4, value="")
            group_total_em = sum(r.total_cost for r in equipment_resources)
            self.ws.cell(row=self.current_row, column=12, value=float(group_total_em))
            self._apply_style(self.ws.cell(row=self.current_row, column=2), self.STYLE_RESOURCE)
            self._apply_style(self.ws.cell(row=self.current_row, column=12), self.STYLE_RESOURCE)
            self.current_row += 1

            self.ws.cell(row=self.current_row, column=3, value="ОТм(ЗТм)")
            self.ws.cell(row=self.current_row, column=4, value="чел.-ч")
            group_total_otm = sum(r.total_cost for r in operator_resources)
            self.ws.cell(row=self.current_row, column=12, value=float(group_total_otm))
            self._apply_style(self.ws.cell(row=self.current_row, column=3), self.STYLE_REGULAR)
            self._apply_style(self.ws.cell(row=self.current_row, column=12), self.STYLE_RESOURCE)
            self.current_row += 1

            for res in item.resources:
                if res.resource_type in [ResourceType.EQUIPMENT, ResourceType.LABOR_OPERATOR]:
                    self._write_resource_detail(res)

            # Группа М (4. М)
            material_resources = [r for r in item.resources if r.resource_type == ResourceType.MATERIAL]

            if material_resources:
                self.ws.cell(row=self.current_row, column=2, value="4")
                self.ws.cell(row=self.current_row, column=3, value="М")
                self.ws.cell(row=self.current_row, column=4, value="")
                group_total_m = sum(r.total_cost for r in material_resources)
                self.ws.cell(row=self.current_row, column=12, value=float(group_total_m))
                self._apply_style(self.ws.cell(row=self.current_row, column=2), self.STYLE_RESOURCE)
                self._apply_style(self.ws.cell(row=self.current_row, column=12), self.STYLE_RESOURCE)
                self.current_row += 1

                for res in material_resources:
                    self._write_resource_detail(res)

    def _write_resource_detail(self, res: Resource) -> None:
        """Запись детальной строки ресурса"""
        SERVICE_CODES = {'1', '2', '4', 'ОТм(ЗТм)'}

        self.ws.cell(row=self.current_row, column=2, value=res.code)
        self.ws.cell(row=self.current_row, column=3, value=res.name)
        self.ws.cell(row=self.current_row, column=4, value=res.unit)

        val = self._format_number(res.quantity_per_unit)
        if val is not None:
            self.ws.cell(row=self.current_row, column=5, value=val)

        val = self._format_number(res.coefficient)
        if val is not None:
            self.ws.cell(row=self.current_row, column=6, value=val)

        val = self._format_number(res.quantity_total)
        if val is not None:
            self.ws.cell(row=self.current_row, column=7, value=val)

        if res.code not in SERVICE_CODES:
            if res.base_price:
                val = self._format_number(res.base_price)
                if val is not None:
                    self.ws.cell(row=self.current_row, column=8, value=val)

            val = self._format_number(res.index)
            if val is not None:
                self.ws.cell(row=self.current_row, column=9, value=val)

            val = self._format_number(res.current_price)
            if val is not None:
                self.ws.cell(row=self.current_row, column=10, value=val)

        val = self._format_number(res.total_cost)
        if val is not None:
            self.ws.cell(row=self.current_row, column=12, value=val)

        for col in range(2, 13):
            self._apply_style(self.ws.cell(row=self.current_row, column=col), self.STYLE_RESOURCE)

        self.current_row += 1

    def _write_work_item_totals(self, item: WorkItem) -> None:
        """Итоговые строки после ГЭСН"""

        # 1. Итого прямые затраты (объединяем только колонки 3-11)
        self._merge_cells(3, 11, self.current_row, self.current_row)
        cell = self.ws.cell(row=self.current_row, column=3, value="Итого прямые затраты")
        self._apply_style(cell, self.STYLE_SUBTOTAL)
        sum_cell = self.ws.cell(row=self.current_row, column=12, value=float(item.direct_costs))
        self._apply_style(sum_cell, self.STYLE_TOTAL)
        self.current_row += 1

        # 2. ФОТ (объединяем только колонки 3-11)
        self._merge_cells(3, 11, self.current_row, self.current_row)
        cell = self.ws.cell(row=self.current_row, column=3, value="ФОТ")
        self._apply_style(cell, self.STYLE_REGULAR)
        sum_cell = self.ws.cell(row=self.current_row, column=12, value=float(item.fot))
        self._apply_style(sum_cell, self.STYLE_RESOURCE)
        self.current_row += 1

        # 3. НР (накладные расходы)
        if hasattr(item, 'base_nr_percent') and item.base_nr_percent:
            nr_amount = item.fot * (item.nr_percent / Decimal("100"))
            work_type_name = getattr(item, 'nr_work_type', item.name) if hasattr(item, 'nr_work_type') else item.name

            nr_code_with_order = f"Пр/812-{item.nr_code}, Приказ № 812/пр от 21.12.2020 п.25"
            self.ws.cell(row=self.current_row, column=2, value=nr_code_with_order)
            self.ws.cell(row=self.current_row, column=3, value=f"НР {work_type_name}")
            self.ws.cell(row=self.current_row, column=4, value="%")

            val = self._format_number(item.base_nr_percent)
            if val is not None:
                self.ws.cell(row=self.current_row, column=5, value=val)

            if hasattr(item, 'nr_reduction') and item.nr_reduction:
                val = self._format_number(item.nr_reduction)
                if val is not None:
                    self.ws.cell(row=self.current_row, column=6, value=val)

            val = self._format_number(item.nr_percent)
            if val is not None:
                self.ws.cell(row=self.current_row, column=7, value=val)

            sum_cell = self.ws.cell(row=self.current_row, column=12, value=float(nr_amount))
            self._apply_style(sum_cell, self.STYLE_RESOURCE)

            for col in range(2, 13):
                self._apply_style(self.ws.cell(row=self.current_row, column=col), self.STYLE_REGULAR)

            self.current_row += 1

        # 4. СП (сметная прибыль)
        if hasattr(item, 'base_sp_percent') and item.base_sp_percent:
            sp_amount = item.fot * (item.sp_percent / Decimal("100"))
            work_type_name = getattr(item, 'sp_work_type', item.name) if hasattr(item, 'sp_work_type') else item.name

            sp_code_with_order = f"Пр/774-{item.sp_code}, Приказ № 774/пр от 11.12.2020 п.16"
            self.ws.cell(row=self.current_row, column=2, value=sp_code_with_order)
            self.ws.cell(row=self.current_row, column=3, value=f"СП {work_type_name}")
            self.ws.cell(row=self.current_row, column=4, value="%")

            val = self._format_number(item.base_sp_percent)
            if val is not None:
                self.ws.cell(row=self.current_row, column=5, value=val)

            if hasattr(item, 'sp_reduction') and item.sp_reduction:
                val = self._format_number(item.sp_reduction)
                if val is not None:
                    self.ws.cell(row=self.current_row, column=6, value=val)

            val = self._format_number(item.sp_percent)
            if val is not None:
                self.ws.cell(row=self.current_row, column=7, value=val)

            sum_cell = self.ws.cell(row=self.current_row, column=12, value=float(sp_amount))
            self._apply_style(sum_cell, self.STYLE_RESOURCE)

            for col in range(2, 13):
                self._apply_style(self.ws.cell(row=self.current_row, column=col), self.STYLE_REGULAR)

            self.current_row += 1

        # 5. Всего по позиции
        total = item.direct_costs
        if hasattr(item, 'nr_percent') and item.nr_percent:
            total += item.fot * (item.nr_percent / Decimal("100"))
        if hasattr(item, 'sp_percent') and item.sp_percent:
            total += item.fot * (item.sp_percent / Decimal("100"))

        # Объединяем колонки 2-3 (только для текста), колонка 12 остаётся отдельно
        cell = self.ws.cell(row=self.current_row, column=3, value="Всего по позиции")
        self._apply_style(cell, self.STYLE_TOTAL)
        sum_cell = self.ws.cell(row=self.current_row, column=12, value=float(total))
        self._apply_style(sum_cell, self.STYLE_TOTAL)
        self.current_row += 1

    def _write_abstract_resource_item(self, abs_res: AbstractResource) -> None:
        """Запись абстрактного ресурса как отдельной позиции (ФСБЦ)"""

        self.ws.cell(row=self.current_row, column=1, value=abs_res.row_num)
        self.ws.cell(row=self.current_row, column=2, value=f"ФСБЦ-{abs_res.code}")
        self.ws.cell(row=self.current_row, column=3, value=abs_res.name)
        self.ws.cell(row=self.current_row, column=4, value=abs_res.unit)

        # Количество (уже умноженное на объём работ)
        self.ws.cell(row=self.current_row, column=5, value=float(abs_res.quantity))
        self.ws.cell(row=self.current_row, column=6, value="1")
        self.ws.cell(row=self.current_row, column=7, value=float(abs_res.quantity))

        if abs_res.base_price:
            self.ws.cell(row=self.current_row, column=8, value=float(abs_res.base_price))

        val = self._format_number(abs_res.index)
        if val is not None:
            self.ws.cell(row=self.current_row, column=9, value=val)

        self.ws.cell(row=self.current_row, column=10, value=float(abs_res.current_price))
        self.ws.cell(row=self.current_row, column=12, value=float(abs_res.total_cost))

        for col in range(1, 13):
            self._apply_style(self.ws.cell(row=self.current_row, column=col), self.STYLE_WORK_MAIN)
        self.current_row += 1

        cell = self.ws.cell(row=self.current_row, column=3, value="Всего по позиции")
        self._apply_style(cell, self.STYLE_TOTAL)
        self.ws.cell(row=self.current_row, column=12, value=float(abs_res.total_cost))
        self._apply_style(self.ws.cell(row=self.current_row, column=12), self.STYLE_TOTAL)
        self.current_row += 1

    def _setup_column_widths(self) -> None:
        for col, width in self.COLUMN_WIDTHS.items():
            col_letter = get_column_letter(col)
            self.ws.column_dimensions[col_letter].width = width

    def export(self, filepath: str) -> str:
        self._write_metadata_header()
        self._write_summary_block()
        self.current_row += 1
        self._write_table_header()

        for section in self.estimate.sections:
            self._write_section(section)
            self.current_row += 1

        self._setup_column_widths()

        os.makedirs(os.path.dirname(filepath) if os.path.dirname(filepath) else '.', exist_ok=True)
        self.wb.save(filepath)

        return filepath
